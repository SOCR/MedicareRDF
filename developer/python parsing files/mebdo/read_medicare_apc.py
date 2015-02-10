"""
    <!-- http://purl.obolibrary.org/obo/MEBDO_0000035 -->

    <owl:Class rdf:about="&obo;MEBDO_0000035">
        <rdfs:label xml:lang="en">Level I Photochemotherapy</rdfs:label>
        <rdfs:subClassOf rdf:resource="&obo;MEBDO_0000030"/>
        <obo:MEBDO_0000034 xml:lang="en">0001</obo:MEBDO_0000034>
    </owl:Class>
"""

from openpyxl import load_workbook

wb = load_workbook(filename = '2013-OPPS-APC-Offset-file.xlsx')

ws = wb.get_sheet_by_name( name = 'Sheet1')

for i in range(4, 373):
    var1 = i
    varID = 32 + i
    print var1
    
    if varID >=100 :
        indexString = '0000' + str(varID) 
    else: 
        indexString = '00000' + str(varID) 
    
    updateFile = open('mdbdo_APC.owl','a')
    
    updateFile.write('    <!-- http://purl.obolibrary.org/obo/MEBDO_' + indexString)
    updateFile.write(' -->')
    updateFile.write('\n')
    updateFile.write('\n')
    updateFile.write('    <owl:Class rdf:about="&obo;MEBDO_' + indexString)
    updateFile.write('">\n')
    
    className = ws.cell('B'+ str(i)).value
    className = str(className)
    updateFile.write('        <rdfs:label xml:lang="en">' + className)
    updateFile.write('</rdfs:label>\n')
    updateFile.write('        <rdfs:subClassOf rdf:resource="&obo;MEBDO_0000030"/>\n')
    
    APCcode= ws.cell('A'+ str(i)).value
    APCcode= str(APCcode)
    updateFile.write('        <obo:MEBDO_0000034 xml:lang="en">' + APCcode)
    updateFile.write('</obo:MEBDO_0000034>\n')
    updateFile.write('    </owl:Class>\n\n\n')
    updateFile.close()