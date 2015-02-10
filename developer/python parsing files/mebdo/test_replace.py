from openpyxl import load_workbook

wb = load_workbook(filename = '2013-OPPS-APC-Offset-file.xlsx')

ws = wb.get_sheet_by_name( name = 'Sheet1')

className = ws.cell('B'+ str(12)).value
className = str(className)
className.replace("&", "&amp;")
print className