"""
   Make output an import file  




    <!-- http://dbpedia.org/resource/United_Sates -->

    <owl:NamedIndividual rdf:about="http://dbpedia.org/resource/United_Sates">
        <rdf:type rdf:resource="&obo;MEBDO_0000020"/>
        <rdfs:label xml:lang="en">United_States</rdfs:label>
    </owl:NamedIndividual>
    


    <!-- http://dbpedia.org/resource/Georgia_(U.S._state) -->

    <owl:NamedIndividual rdf:about="http://dbpedia.org/resource/Georgia_(U.S._state)">
        <rdf:type rdf:resource="&obo;MEBDO_0000011"/>
        <rdfs:label xml:lang="en">Georgia_(U.S._state)</rdfs:label>
        <obo:BFO_0000176 rdf:resource="http://dbpedia.org/resource/United_Sates"/>
    </owl:NamedIndividual>
    


    <!-- http://dbpedia.org/resource/Hephzibah,_Georgia -->

    <owl:NamedIndividual rdf:about="http://dbpedia.org/resource/Hephzibah,_Georgia">
        <rdf:type rdf:resource="&obo;MEBDO_0000010"/>
        <rdfs:label xml:lang="en">Hephzibah,_Georgia</rdfs:label>
		<obo:IAO_0000232>city-zip statement was extracted from DBpedia at 01/28/2015</obo:IAO_0000232>
        <obo:BFO_0000176 rdf:resource="http://dbpedia.org/resource/Richmond_County,_Georgia"/>
        
        <obo:MEBDO_0000410 rdf:resource="&obo;MEBDO_0000412"/>
    </owl:NamedIndividual>
    


    <!-- http://dbpedia.org/resource/Richmond_County,_Georgia -->

    <owl:NamedIndividual rdf:about="http://dbpedia.org/resource/Richmond_County,_Georgia">
        <rdf:type rdf:resource="&obo;MEBDO_0000015"/>
        <rdfs:label xml:lang="en">Richmond_County,_Georgia</rdfs:label>
        <obo:BFO_0000176 rdf:resource="http://dbpedia.org/resource/Georgia_(U.S._state)"/>
    </owl:NamedIndividual>
    


    <!-- http://purl.obolibrary.org/obo/MEBDO_0000412 -->

    <owl:NamedIndividual rdf:about="&obo;MEBDO_0000412">
        <rdf:type rdf:resource="&obo;MEBDO_0000012"/>
        <rdfs:label xml:lang="en">30815 zip code zone in 2012</rdfs:label>
        <rdfs:seeAlso rdf:datatype="&xsd;anyURI">http://zipcode.org/30815</rdfs:seeAlso>
    </owl:NamedIndividual>

"""
## policy for zipcode id: 947+++++


from openpyxl import load_workbook
import re
from zipcodewriteFile import write_zipcode


## getting object property and class list
annotationpropertyList = ['IAO_0000232']
objectpropertyList = ['MEBDO_0000410','BFO_0000176']
classList = ['MEBDO_0000020','MEBDO_0000010','MEBDO_0000011','MEBDO_0000012', 'MEBDO_0000015']

fileName = 'mebdo_city_county_import.owl'

from writehead import write_head
write_head(annotationpropertyList, objectpropertyList, classList, fileName)

wb = load_workbook(filename = 'US_state_city_county_postalCode.xlsx')
ws = wb.get_sheet_by_name(name = 'US_state_city_county_postalCode')

for i in range(2,7450): ##2,7450
    print i
    print "starts...."
    cityURI = ws.cell('A'+ str(i)).value
    cityName = re.sub('http://dbpedia.org/resource/', "", str(cityURI))
    countyURI = ws.cell('B'+ str(i)).value
    countyName = re.sub('http://dbpedia.org/resource/', "", str(countyURI))
    stateURI = ws.cell('C'+ str(i)).value
    stateName = re.sub('http://dbpedia.org/resource/', "", str(stateURI))
    postalCodeString = ws.cell('D'+ str(i)).value 
    curatorNote = ws.cell('E'+ str(i)).value
    
    ##open file
    updateFile =  open('mebdo_city_county_import.owl','a')
    
    ## individual city without notion of zipcode
    updateFile.write('    <!-- ' + str(cityURI) + ' -->\n')
    updateFile.write('    <owl:NamedIndividual rdf:about="'+str(cityURI)+'">\n')
    updateFile.write('        <rdf:type rdf:resource="http://purl.obolibrary.org/obo/MEBDO_0000010"/>\n')
    ## add curation note
    curatorNote = str(curatorNote)
    if curatorNote <> 'None':
        if re.search('not', curatorNote, re.IGNORECASE):
            comment = curatorNote.split(';')
            if len(comment) >1 :
                updateFile.write('        <obo:IAO_0000232>'+ comment[1] +'</obo:IAO_0000232>\n')
                updateFile.write('        <obo:IAO_0000232>city was not found in zipcode.org on 01/28/2015</obo:IAO_0000232>\n')
            else :
                updateFile.write('        <obo:IAO_0000232>city was not found in zipcode.org on 01/28/2015</obo:IAO_0000232>\n')
        elif re.search('zipcode.org', curatorNote):
            updateFile.write('        <obo:IAO_0000232>city-zip statement was extracted from zipcode.org during 01/27/2015-01/29/2015</obo:IAO_0000232>\n')
        else :
            updateFile.write('        <obo:IAO_0000232>'+ curatorNote +'</obo:IAO_0000232>\n')
    else:
        updateFile.write('    <obo:IAO_0000232>city-zip statement was extracted from DBpedia on 01/27/2015</obo:IAO_0000232>\n')
    updateFile.write('        <rdfs:label xml:lang="en">' + cityName + '</rdfs:label>\n')
    updateFile.write('        <obo:BFO_0000176 rdf:resource="' + countyURI + '"/>\n')
    updateFile.write('        <obo:BFO_0000176 rdf:resource="'+stateURI+'"/>\n    </owl:NamedIndividual>\n\n\n')

    ##individual county
    updateFile.write('    <!-- http://dbpedia.org/resource/' +countyName +' -->\n')
    updateFile.write('    <owl:NamedIndividual rdf:about="' +countyURI+'">\n')
    updateFile.write('        <rdf:type rdf:resource="http://purl.obolibrary.org/obo/MEBDO_0000015"/>\n        <rdfs:label xml:lang="en">'+countyName+'</rdfs:label>\n    </owl:NamedIndividual>\n\n\n')

    ##individual state
    updateFile.write('    <!-- '+stateURI+' -->\n')
    updateFile.write('    <owl:NamedIndividual rdf:about="'+stateURI+'">\n')
    updateFile.write('        <rdf:type rdf:resource="http://purl.obolibrary.org/obo/MEBDO_0000011"/>\n        <rdfs:label xml:lang="en">'+stateName+'</rdfs:label>\n')
    updateFile.write('        <obo:BFO_0000176 rdf:resource="http://dbpedia.org/resource/United_Sates"/>\n    </owl:NamedIndividual>\n\n\n')    

    
    ## zipcode 
    if postalCodeString : 
        zipcode = str(postalCodeString)
        searchObj0 = re.search('-*,',zipcode)
        searchObj = re.search(',',zipcode)
        searchObj2 = re.search('-',zipcode)
        if searchObj0:
            zipcodelist = zipcode.split(',')
            y = len(zipcodelist)
            for s in range(0,y):
                zipcode_list = zipcodelist[s]
                zipcode_list_list = zipcode_list.split('-')
                if len(zipcode_list_list) > 1 :
                    m0 = int(zipcode_list_list[1]) - int(zipcode_list_list[0]) + 1
                    for r0 in range(0, m0):	
                        zipcodeF = int(zipcode_list_list[0]) + r0
                        zipcodeF = str(zipcodeF)
                        write_zipcode(zipcodeF,cityURI)
                else:
                    zipcodeF = zipcode_list_list[0]
                    write_zipcode(zipcodeF,cityURI)
        elif searchObj:
            zipcodelist = zipcode.split(',')
            x = len(zipcodelist)
            for n in range(0,x):
                zipcodeF = zipcodelist[n]
                zipcodeF = zipcodeF.strip()
                write_zipcode(zipcodeF,cityURI)
        elif searchObj2:
            zipcodelist2 = zipcode.split('-')
            m=int(zipcodelist2[1]) - int(zipcodelist2[0]) + 1
            for r in range(0,m):
                zipcodeF = int(zipcodelist2[0]) + r
                zipcodeF = str(zipcodeF)
                write_zipcode(zipcodeF,cityURI)
        else:
            zipcodeF = zipcode.strip()
            write_zipcode(zipcodeF,cityURI)		    
     
    
    print "finished!\n"

updateFile.write('</rdf:RDF>')
updateFile.close()


"""
print ''.join(open('C:\Users\yuln\Documents\mebdo\developer\city and states\zipcode.txt').read().splitlines())
"""