""" county is part of state

    <!-- http://dbpedia.org/resource/Richmond_County,_Georgia -->

    <owl:NamedIndividual rdf:about="http://dbpedia.org/resource/Richmond_County,_Georgia">
        <rdf:type rdf:resource="&obo;MEBDO_0000015"/>
        <rdfs:label xml:lang="en">Richmond_County,_Georgia</rdfs:label>
        <obo:BFO_0000176 rdf:resource="http://dbpedia.org/resource/Georgia_(U.S._state)"/>
    </owl:NamedIndividual>
	
"""

from openpyxl import load_workbook
import re

annotationpropertyList = ['IAO_0000232']
objectpropertyList = ['BFO_0000176']
classList = []

fileName = 'mebdo_county_state_import.owl'

from writehead import write_head
write_head(annotationpropertyList, objectpropertyList, classList, fileName)

wb = load_workbook(filename = 'US_county_state.xlsx')
ws = wb.get_sheet_by_name(name = 'Sheet1')



for i in range(2,2695):
    print i
    print "starts...."
    countyURI = ws.cell('A'+ str(i)).value
    countyName = re.sub('http://dbpedia.org/resource/', "", str(countyURI))
    stateURI = ws.cell('B'+ str(i)).value

    ## open file
    updateFile =  open(fileName,'a')

    ##individual county
    updateFile.write('    <!-- http://dbpedia.org/resource/' +countyName +' -->\n')
    updateFile.write('    <owl:NamedIndividual rdf:about="' +countyURI+'">\n')
    updateFile.write('    <obo:IAO_0000232>county-state statement was extracted from DBpedia on 01/29/2015</obo:IAO_0000232>\n')
    updateFile.write('        <obo:BFO_0000176 rdf:resource="'+stateURI+'"/>\n    </owl:NamedIndividual>\n\n\n') 

    print "finished!\n"

updateFile.write('</rdf:RDF>')
updateFile.close()